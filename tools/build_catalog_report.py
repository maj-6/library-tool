"""Build a CH Library status report spreadsheet.

Reads ch_library.xlsx and writes output/ch_library_report.xlsx with the same
columns plus four added columns:
  - In WHL             (offline; matched against whl_catalog.csv)
  - Available online   (opt-in --online; Internet Archive search, cached)
  - In local library   (offline; matched against the local book set)
  - Copyright status   (offline; via copyright_renewals.csv)

The copyright, In WHL, and local-library columns are computed offline and are
always populated. Only 'Available online' needs the network; it defaults to
"not checked" unless --online is passed, and is rate-limited via --limit with
on-disk caching so runs are resumable.

The matching logic and database loaders live in tools/catalog_checks.py,
shared with the review web app's per-submission checks.

Run with python3:
  python3 tools/build_catalog_report.py              # offline columns only
  python3 tools/build_catalog_report.py --online --limit 200
"""
from __future__ import annotations

import argparse
import json
import sys
import time
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
import catalog_checks as checks  # noqa: E402
import libcommon as lib  # noqa: E402
import whl_client as whl  # noqa: E402

REPORT_PATH = lib.OUTPUT_DIR / "ch_library_report.xlsx"
LOCAL_PARTIAL_PATH = lib.ROOT / "local_library_partial.json"
ONLINE_CACHE = lib.OUTPUT_DIR / ".online_cache.json"

ADDED_COLUMNS = ["In WHL", "Available online", "In local library", "Copyright status"]

# Cached "yes"/"no" answers are stable and kept forever. A cached "error" is a
# transient network failure, so it is only trusted for this long before the next
# run is allowed to retry it (older bare-string caches, which have no timestamp,
# are always retried).
ERROR_RETRY_TTL = 6 * 3600  # seconds


# --- CH Library title/author helpers ---------------------------------------

def ch_title(row: dict) -> str:
    return str(row.get("PUBLICATION", "") or "").replace("_", " ").strip()


def copyright_status(row: dict, ren: dict, this_year: int) -> str:
    return checks.copyright_status_for(
        ch_title(row), str(row.get("AUTHORS", "") or ""), row.get("YEAR_OF_PU"),
        ren, this_year,
    )


# --- local library ----------------------------------------------------------

def load_local_library() -> list[dict]:
    """Local book set: the hand-entered manual entries + optional partial."""
    books: list[dict] = []

    def add(title, author):
        title = str(title or "").strip()
        author = str(author or "").strip()
        if title or author:
            books.append({"title": title, "author": author})

    manual = lib.load_json(lib.MANUAL_ENTRIES_PATH, {})
    for entry in (manual.values() if isinstance(manual, dict) else manual):
        add(entry.get("title"), entry.get("author"))
    partial = lib.load_json(LOCAL_PARTIAL_PATH, [])
    for entry in (partial if isinstance(partial, list) else []):
        if isinstance(entry, dict):
            add(entry.get("title"), entry.get("author") or entry.get("author_last"))
    return books


def in_local_library(row: dict, local: list[dict]) -> str:
    title = ch_title(row)
    author = str(row.get("AUTHORS", "") or "")
    for b in local:
        if checks.title_author_match(title, author, b["title"], b["author"]):
            return "yes"
    return "no"


# --- online availability (opt-in) -------------------------------------------

def _cache_entry(value) -> tuple[str, float | None]:
    """Normalise a cached value to ``(status, ts)``.

    New entries are ``{"status": ..., "ts": ...}`` dicts; legacy caches stored a
    bare ``"yes"``/``"no"``/``"error"`` string with no timestamp. Anything
    unrecognised coerces to a stale error so it will be re-checked.
    """
    if isinstance(value, dict):
        status = value.get("status")
        ts = value.get("ts")
        if status in ("yes", "no", "error"):
            return status, (ts if isinstance(ts, (int, float)) else None)
        return "error", None
    if value in ("yes", "no", "error"):
        return value, None
    return "error", None


def check_online(title: str, author: str, cache: dict, allow_request: bool = True) -> tuple[str, bool]:
    """Resolve Internet Archive availability for a book.

    Returns ``(status, did_request)`` where ``status`` is ``"yes"``/``"no"``/
    ``"error"``/``"not checked"`` and ``did_request`` is True only when a real
    network call was made (the caller uses this to spend rate-limit budget and
    sleep). A cached ``yes``/``no`` — or a cached ``error`` still inside
    ``ERROR_RETRY_TTL`` — resolves for free and never spends budget, so a resumed
    run advances past already-checked rows instead of re-spending on them. When
    ``allow_request`` is False and the answer is not cached, returns
    ``("not checked", False)`` without touching the network.
    """
    key = whl._normalize(title) + "|" + whl._normalize(author)
    if key in cache:
        status, ts = _cache_entry(cache[key])
        if status in ("yes", "no"):
            return status, False
        # A cached error is only trusted for a while, and legacy bare errors
        # (no timestamp) are always retried; otherwise fall through to re-check.
        if ts is not None and (time.time() - ts) < ERROR_RETRY_TTL:
            return "error", False
    if not allow_request:
        return "not checked", False
    words = " ".join(whl._normalize(title).split()[:8])
    q = f'title:({words}) AND mediatype:texts'
    if author:
        q += f' AND creator:({whl._normalize(author)})'
    url = "https://archive.org/advancedsearch.php?" + urllib.parse.urlencode(
        {"q": q, "rows": 0, "output": "json"}
    )
    try:
        req = urllib.request.Request(url, headers={"User-Agent": whl.USER_AGENT})
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        found = int(data.get("response", {}).get("numFound", 0))
        value = "yes" if found > 0 else "no"
    except Exception:
        value = "error"
    cache[key] = {"status": value, "ts": time.time()}
    return value, True


def resolve_online(pairs, cache, *, limit, sleep_s=0.0, save=None) -> list[str]:
    """Resolve online availability for a sequence of ``(title, author)`` pairs.

    Spends at most ``limit`` network requests (``limit == 0`` means unlimited),
    counting only genuine requests — cached rows resolve for free and never cost
    budget or a sleep. After each real request the cache is flushed via ``save``
    (if given) so an interrupted run keeps every lookup it completed. Returns the
    per-pair status list in order.
    """
    budget = limit
    out: list[str] = []
    for title, author in pairs:
        allow = (limit == 0 or budget > 0)
        status, did_request = check_online(title, author, cache, allow_request=allow)
        if did_request:
            budget -= 1
            if save is not None:
                save(cache)
            if sleep_s:
                time.sleep(sleep_s)
        out.append(status)
    return out


# --- report build -----------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Build the CH Library status report.")
    parser.add_argument("--online", action="store_true", help="Check Internet Archive availability (network).")
    parser.add_argument("--limit", type=int, default=200,
                        help="Max Internet Archive lookups (0 = no limit). Default 200.")
    parser.add_argument("--sleep", type=float, default=0.2, help="Delay between network calls, seconds.")
    args = parser.parse_args()

    if not lib.XLSX_PATH.exists():
        parser.error(f"Missing {lib.XLSX_PATH}")

    this_year = datetime.now().year
    print("loading copyright renewals ...")
    ren = checks.load_renewals()
    print(f"  {len(ren['entries'])} renewal records indexed")
    whlcat = checks.load_whl_catalog()
    print(f"  {len(whlcat['entries'])} WHL catalogue records indexed")
    local = load_local_library()
    print(f"  {len(local)} local library books loaded")

    online_cache = lib.load_json(ONLINE_CACHE, {}) if args.online else {}

    src = openpyxl.load_workbook(lib.XLSX_PATH, read_only=True, data_only=True)
    ws = src.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    headers = [str(h) if h is not None else "" for h in next(rows)]

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = ws.title
    out_ws.append(headers + ADDED_COLUMNS)
    out_ws.freeze_panes = "A2"

    status_counts: dict[str, int] = {}
    whl_counts: dict[str, int] = {}
    pending: list[tuple[list, str, str, str]] = []  # input values + whl/local/copyright flags
    pairs: list[tuple[str, str]] = []               # (title, author) for the online lookup
    n = 0
    for values in rows:
        if values is None or all(v is None for v in values):
            continue
        row = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
        n += 1

        title = ch_title(row)
        author = str(row.get("AUTHORS", "") or "")
        year = whl._year(row.get("YEAR_OF_PU")) or ""

        status = copyright_status(row, ren, this_year)
        bucket = status.split(" (")[0]
        status_counts[bucket] = status_counts.get(bucket, 0) + 1
        local_flag = in_local_library(row, local)
        whl_flag = checks.whl_catalog_flag(title, author, year, whlcat)
        whl_counts[whl_flag] = whl_counts.get(whl_flag, 0) + 1

        pending.append((list(values), whl_flag, local_flag, status))
        pairs.append((title, author))
        if n % 500 == 0:
            print(f"  processed {n} rows ...")

    # Resolve online availability in a single pass: cached rows are free, only
    # genuine lookups spend --limit budget, and the cache is flushed after each
    # real request so an interrupted run keeps every lookup it completed.
    if args.online:
        online_flags = resolve_online(
            pairs, online_cache, limit=args.limit, sleep_s=args.sleep,
            save=lambda c: lib.save_json(ONLINE_CACHE, c),
        )
    else:
        online_flags = ["not checked"] * len(pairs)

    for (values, whl_flag, local_flag, status), online_flag in zip(pending, online_flags):
        out_ws.append(list(values) + [whl_flag, online_flag, local_flag, status])

    lib.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_wb.save(REPORT_PATH)
    if args.online:
        lib.save_json(ONLINE_CACHE, online_cache)

    print(f"\nwrote {REPORT_PATH}")
    print(f"rows: {n} | columns: {len(headers) + len(ADDED_COLUMNS)}")
    print("copyright status distribution:")
    for k in sorted(status_counts, key=lambda k: -status_counts[k]):
        print(f"  {status_counts[k]:5d}  {k}")
    print("In WHL distribution:")
    for k in sorted(whl_counts, key=lambda k: -whl_counts[k]):
        print(f"  {whl_counts[k]:5d}  {k}")


if __name__ == "__main__":
    main()
